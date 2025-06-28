# inventory/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Employee, ITAsset, Department, Position, AssetType, OwnerCompany, AssetHistory, Team, User, Role, Ticket, UserProfile, TeamMember
from .forms import EmployeeForm, ITAssetForm, UserForm, UserProfileForm, RoleForm, TicketForm, AssetTypeForm, DepartmentForm, CompanyForm, TeamForm
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.contrib.auth import authenticate, login, logout
from django.utils import timezone
import json
import time
from django.core.exceptions import PermissionDenied
import logging
from .mixins import PermissionRequiredMixin, SuperuserRequiredMixin

logger = logging.getLogger(__name__)

@login_required
def home(request):
    if not request.user.is_superuser:
        return render(request, 'inventory/home.html', {
            'show_no_data': True
        })
    # Get total counts
    total_assets = ITAsset.objects.count()
    total_employees = Employee.objects.count()
    available_assets = ITAsset.objects.filter(status='available').count()
    assigned_assets = ITAsset.objects.filter(status='assigned').count()
    maintenance_assets = ITAsset.objects.filter(status='maintenance').count()
    retired_assets = ITAsset.objects.filter(status='retired').count()

    # Get asset type distribution
    asset_types = AssetType.objects.all()
    asset_type_distribution = []
    for asset_type in asset_types:
        count = ITAsset.objects.filter(asset_type=asset_type).count()
        asset_type_distribution.append({
            'name': asset_type.display_name,
            'count': count
        })

    # Get owner company distribution
    owner_companies = OwnerCompany.objects.all()
    owner_company_distribution = []
    for company in owner_companies:
        count = ITAsset.objects.filter(owner=company).count()
        owner_company_distribution.append({
            'name': company.name,
            'count': count
        })

    # Get recent assets
    recent_assets = ITAsset.objects.select_related('asset_type', 'owner', 'assigned_to').order_by('-id')[:5]

    # Get assets with expiring warranty (within 3 months)
    today = datetime.now().date()
    three_months_later = today + timedelta(days=90)
    expiring_warranty_assets = ITAsset.objects.filter(
        warranty_expiry__gte=today,
        warranty_expiry__lte=three_months_later
    ).select_related('asset_type', 'owner', 'assigned_to')

    # Get recent asset history
    recent_history = AssetHistory.objects.select_related('asset', 'assigned_to').order_by('-date')[:5]

    context = {
        'total_assets': total_assets,
        'total_employees': total_employees,
        'available_assets': available_assets,
        'assigned_assets': assigned_assets,
        'maintenance_assets': maintenance_assets,
        'retired_assets': retired_assets,
        'asset_type_distribution': asset_type_distribution,
        'owner_company_distribution': owner_company_distribution,
        'recent_assets': recent_assets,
        'expiring_warranty_assets': expiring_warranty_assets,
        'recent_history': recent_history,
    }
    return render(request, 'inventory/home.html', context)

class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'inventory/employee_list.html'
    context_object_name = 'employees'

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_employee'):
            return super().get_queryset()
        return Employee.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.all()
        context['companies'] = OwnerCompany.objects.all()
        access_denied = not self.request.user.has_perm('inventory.view_employee')
        context['access_denied'] = access_denied
        context['can_add_employee'] = self.request.user.has_perm('inventory.add_employee') and not access_denied
        context['can_change_employee'] = self.request.user.has_perm('inventory.change_employee') and not access_denied
        context['can_delete_employee'] = self.request.user.has_perm('inventory.delete_employee') and not access_denied
        return context

class EmployeeDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Employee
    template_name = 'inventory/employee_detail.html'
    context_object_name = 'employee'
    permission_required = 'inventory.view_employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_change_employee'] = self.request.user.has_perm('inventory.change_employee')
        context['can_delete_employee'] = self.request.user.has_perm('inventory.delete_employee')
        return context

class EmployeeCreateView(SuperuserRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'inventory/employee_form.html'
    success_url = reverse_lazy('inventory:employee_list')

    def form_valid(self, form):
        messages.success(self.request, 'Employee created successfully.')
        return super().form_valid(form)

class EmployeeUpdateView(SuperuserRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'inventory/employee_form.html'
    success_url = reverse_lazy('inventory:employee_list')

    def form_valid(self, form):
        messages.success(self.request, 'Employee updated successfully.')
        return super().form_valid(form)

class EmployeeDeleteView(SuperuserRequiredMixin, DeleteView):
    model = Employee
    template_name = 'inventory/employee_confirm_delete.html'
    success_url = reverse_lazy('inventory:employee_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Employee deleted successfully.')
        return super().delete(request, *args, **kwargs)

@login_required
def asset_assign(request):
    """View for assigning assets to employees."""
    if request.method == 'POST':
        asset_id = request.POST.get('asset')
        employee_id = request.POST.get('employee')
        
        try:
            asset = ITAsset.objects.get(pk=asset_id)
            employee = Employee.objects.get(pk=employee_id)
            
            asset.assigned_to = employee
            asset.status = 'assigned'
            asset.save()
            
            messages.success(request, f'Asset "{asset.name}" has been assigned to {employee.get_full_name()}.')
            return redirect('inventory:employee_detail', pk=employee.pk)
        except (ITAsset.DoesNotExist, Employee.DoesNotExist):
            messages.error(request, 'Invalid asset or employee selected.')
            return redirect('inventory:employee_list')
    
    # Get the employee from the query parameter
    employee_id = request.GET.get('employee')
    if employee_id:
        try:
            employee = Employee.objects.get(pk=employee_id)
        except Employee.DoesNotExist:
            messages.error(request, 'Employee not found.')
            return redirect('inventory:employee_list')
    else:
        employee = None
    
    # Get available assets
    available_assets = ITAsset.objects.filter(status='available')
    
    context = {
        'employee': employee,
        'available_assets': available_assets,
    }
    
    return render(request, 'inventory/asset_assign.html', context)

class ITAssetListView(LoginRequiredMixin, ListView):
    model = ITAsset
    template_name = 'inventory/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 10

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_asset'):
            return super().get_queryset()
        return ITAsset.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['asset_types'] = AssetType.objects.all()
        context['status_choices'] = ITAsset.STATUS_CHOICES
        context['manufacturers'] = ITAsset.objects.exclude(manufacturer='').values_list('manufacturer', flat=True).distinct()
        access_denied = not self.request.user.has_perm('inventory.view_asset')
        context['access_denied'] = access_denied
        context['can_add_asset'] = self.request.user.has_perm('inventory.add_asset') and not access_denied
        context['can_change_asset'] = self.request.user.has_perm('inventory.change_asset') and not access_denied
        context['can_delete_asset'] = self.request.user.has_perm('inventory.delete_asset') and not access_denied
        context['can_download_asset'] = self.request.user.has_perm('inventory.download_asset') and not access_denied
        context['can_upload_asset'] = self.request.user.has_perm('inventory.upload_asset') and not access_denied
        return context

class ITAssetDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ITAsset
    template_name = 'inventory/asset_detail.html'
    context_object_name = 'asset'
    permission_required = 'inventory.view_asset'

class ITAssetCreateView(SuperuserRequiredMixin, CreateView):
    model = ITAsset
    form_class = ITAssetForm
    template_name = 'inventory/asset_form.html'
    success_url = reverse_lazy('inventory:asset_list')

    def form_valid(self, form):
        messages.success(self.request, 'IT Asset created successfully.')
        return super().form_valid(form)

class ITAssetUpdateView(SuperuserRequiredMixin, CreateView):
    model = ITAsset
    form_class = ITAssetForm
    template_name = 'inventory/asset_form.html'
    success_url = reverse_lazy('inventory:asset_list')

    def form_valid(self, form):
        messages.success(self.request, 'IT Asset updated successfully.')
        return super().form_valid(form)

class ITAssetDeleteView(SuperuserRequiredMixin, DeleteView):
    model = ITAsset
    template_name = 'inventory/asset_confirm_delete.html'
    success_url = reverse_lazy('inventory:asset_list')

@login_required
def employee_upload(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # Skip header row
                for row in ws.iter_rows(min_row=2):
                    try:
                        # Get or create department first
                        department_name = row[5].value
                        position_name = row[6].value

                        if department_name:
                            department, _ = Department.objects.get_or_create(name=department_name)
                            
                            # Get or create position object for the Position model, but don't use the object for the employee's position field
                            if position_name:
                                Position.objects.get_or_create(
                                    name=position_name,
                                    department=department
                                )
                        else:
                            department = None
                        
                        # Get company
                        company_name = row[7].value
                        try:
                            company = OwnerCompany.objects.get(name=company_name)
                        except OwnerCompany.DoesNotExist:
                            messages.error(request, f'Row {row[0].row}: Company "{company_name}" not found. Please create it first.')
                            continue
                        
                        # Get employment status and related dates
                        employment_status_original = row[9].value
                        employment_status = employment_status_original.lower() if isinstance(employment_status_original, str) else employment_status_original
                        
                        retirement_date = row[10].value
                        resignation_date = row[11].value
                        training_start_date = row[12].value
                        training_end_date = row[13].value

                        # Validate status
                        valid_statuses = [choice[0] for choice in Employee.EMPLOYMENT_STATUS_CHOICES]
                        if employment_status not in valid_statuses:
                            messages.error(request, f'Row {row[0].row}: Invalid employment status "{employment_status_original}".')
                            continue

                        # Create employee
                        employee = Employee(
                            employee_id=row[0].value,
                            national_id=row[1].value,
                            first_name=row[2].value,
                            last_name=row[3].value,
                            email=row[4].value,
                            department=department,
                            position=position_name,
                            company=company,
                            hire_date=row[8].value,
                            employment_status=employment_status,
                            retirement_date=retirement_date if employment_status == 'retired' else None,
                            resignation_date=resignation_date if employment_status == 'resigned' else None,
                            training_start_date=training_start_date if employment_status == 'training' else None,
                            training_end_date=training_end_date if employment_status == 'training' else None,
                        )
                        employee.save()
                    except Exception as e:
                        messages.error(request, f'Error processing row: {str(e)}')
                        continue
                
                messages.success(request, 'Employees uploaded successfully!')
                return redirect('inventory:employee_list')
            except Exception as e:
                messages.error(request, f'Error uploading file: {str(e)}')
    
    return render(request, 'inventory/employee_upload.html')

@login_required
def download_employee_template(request):
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        'Employee ID *',
        'National ID *',
        'First Name *',
        'Last Name *',
        'Email *',
        'Department *',
        'Position *',
        'Company *',
        'Hire Date',
        'Employment Status *',
        'Retirement Date',
        'Resignation Date',
        'Training Start Date',
        'Training End Date'
    ]
    ws.append(headers)
    
    # Add example row
    example = [
        'EMP001',
        '12345678901234',
        'John',
        'Doe',
        'john.doe@example.com',
        'IT',
        'Software Engineer',
        'CTP',
        '2024-01-01',
        'employed', # Default status
        '', # Retirement Date
        '', # Resignation Date
        '', # Training Start Date
        '' # Training End Date
    ]
    ws.append(example)
    
    # Add available departments
    departments = Department.objects.all()
    if departments.exists():
        ws.append([])  # Add empty row
        ws.append(['Available Departments:'])
        for dept in departments:
            ws.append([dept.get_name_display()])
    
    # Add available companies
    companies = OwnerCompany.objects.all()
    if companies.exists():
        ws.append([])  # Add empty row
        ws.append(['Available Companies:'])
        for company in companies:
            ws.append([company.name])
    
    # Add available employment statuses
    ws.append([]) # Add empty row
    ws.append(['Available Employment Statuses:'])
    for status_code, status_name in Employee.EMPLOYMENT_STATUS_CHOICES:
        ws.append([f'{status_code} - {status_name}'])
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Style the sections headers
    if departments.exists():
        dept_header = ws.cell(row=4, column=1)
        dept_header.font = openpyxl.styles.Font(bold=True)
    
    if companies.exists():
        company_header = ws.cell(row=departments.count() + 6, column=1)
        company_header.font = openpyxl.styles.Font(bold=True)
    
    # Add data validation for company column
    company_validation = DataValidation(type="list", formula1=f'"{",".join(companies.values_list("name", flat=True))}"')
    company_validation.add(f'H2:H{len(example) + 1}')
    ws.add_data_validation(company_validation)
    
    # Add data validation for department column
    department_validation = DataValidation(type="list", formula1=f'"{",".join(departments.values_list("name", flat=True))}"')
    department_validation.add(f'F2:F{len(example) + 1}')
    ws.add_data_validation(department_validation)
    
    # Add data validation for employment status column
    status_codes = [status[0] for status in Employee.EMPLOYMENT_STATUS_CHOICES]
    status_validation = DataValidation(type="list", formula1=f'"{",".join(status_codes)}"')
    status_validation.add(f'J2:J{ws.max_row}')
    ws.add_data_validation(status_validation)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
    
    wb.save(response)
    return response

@login_required
def download_employee_data(request):
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        'Employee ID',
        'National ID',
        'First Name',
        'Last Name',
        'Email',
        'Department',
        'Position',
        'Company',
        'Hire Date',
        'Employment Status',
        'Resignation Date',
        'Retirement Date',
        'Training Start Date',
        'Training End Date',
    ]
    ws.append(headers)
    
    # Add employee data
    employees = Employee.objects.all()
    for employee in employees:
        row = [
            employee.employee_id,
            employee.national_id,
            employee.first_name,
            employee.last_name,
            employee.email,
            str(employee.department) if employee.department else '',
            str(employee.position) if employee.position else '',
            str(employee.company.name) if employee.company else '',  # Convert company to string
            employee.hire_date,
            employee.get_employment_status_display(),
            employee.resignation_date,
            employee.retirement_date,
            employee.training_start_date,
            employee.training_end_date,
        ]
        ws.append(row)
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
    
    wb.save(response)
    return response

@login_required
def download_asset_template(request):
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        'Asset Name *',
        'Asset Type *',
        'Serial Number *',
        'Model *',
        'Manufacturer',
        'Owner Company *',
        'Assigned Employee ID',
        'Wi-Fi MAC Address',
        'Ethernet MAC Address *',
        'Delivery Letter Code',
        'Purchase Date',
        'Receipt Date',
        'Warranty Expiry',
        'Status *'  # Added status field
    ]
    ws.append(headers)
    
    # Add example row
    example = [
        'Laptop-001',
        'laptop',  # Use the internal name, not display name
        'SN123456',
        'ThinkPad X1',
        'Lenovo',
        'CTP',  # Company name
        'EMP001',  # Employee ID
        '00:1A:2B:3C:4D:5E',
        '00:1A:2B:3C:4D:5F',
        'DL-2024-001',
        '2024-01-01',
        '2024-01-15',
        '2027-01-01',
        'available'  # Default status
    ]
    ws.append(example)
    
    # Add available asset types
    ws.append([])  # Add empty row
    ws.append(['Available Asset Types:'])
    ws.append(['Internal Name - Display Name'])
    for asset_type in AssetType.objects.all().order_by('name'):
        ws.append([f'{asset_type.name} - {asset_type.display_name}'])
    
    # Add available statuses
    ws.append([])  # Add empty row
    ws.append(['Available Statuses:'])
    ws.append(['Internal Name - Display Name'])
    for status_code, status_name in ITAsset.STATUS_CHOICES:
        ws.append([f'{status_code} - {status_name}'])
    
    # Add owner companies section
    ws.append([])  # Add empty row
    ws.append(['Owner Companies:'])
    for company in OwnerCompany.objects.all().order_by('name'):
        ws.append([company.name])
    
    # Add available employees section
    ws.append([])  # Add empty row
    ws.append(['Available Employees:'])
    ws.append(['Format: Employee ID - Full Name (Company)'])
    ws.append([])
    
    # Group employees by company
    for company in OwnerCompany.objects.all():
        ws.append([f'--- {company.name} Employees ---'])
        employees = Employee.objects.filter(
            employment_status=Employee.STATUS_EMPLOYED,
            company=company
        ).order_by('employee_id')
        
        for employee in employees:
            ws.append([f"{employee.employee_id} - {employee.get_full_name()} ({company.name})"])
        ws.append([])  # Add empty row between companies
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Style the sections headers
    ws.cell(row=4, column=1).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=5, column=1).font = openpyxl.styles.Font(italic=True)  # Style the "Internal Name - Display Name" row
    
    # Find and style other section headers
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value in ['Owner Companies:', 'Available Employees:', 'Available Statuses:'] or \
           (cell.value and isinstance(cell.value, str) and cell.value.startswith('---')):
            cell.font = openpyxl.styles.Font(bold=True)
    
    # Add data validation for asset type column
    asset_types = AssetType.objects.all()
    if asset_types.exists():
        asset_type_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(asset_types.values_list("name", flat=True))}"'
        )
        asset_type_validation.add(f'B2:B{len(example) + 1}')
        ws.add_data_validation(asset_type_validation)
    
    # Add data validation for company column
    companies = OwnerCompany.objects.all()
    if companies.exists():
        company_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(companies.values_list("name", flat=True))}"'
        )
        company_validation.add(f'F2:F{len(example) + 1}')
        ws.add_data_validation(company_validation)
    
    # Add data validation for status column
    status_validation = DataValidation(
        type="list",
        formula1=f'"{",".join([status[0] for status in ITAsset.STATUS_CHOICES])}"'
    )
    status_validation.add(f'N2:N{len(example) + 1}')
    ws.add_data_validation(status_validation)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asset_template.xlsx'
    
    wb.save(response)
    return response

@login_required
def asset_upload(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # Skip header row
                for row in ws.iter_rows(min_row=2):
                    try:
                        # Check required fields
                        required_fields = {
                            'name': row[0].value,
                            'asset_type': row[1].value,
                            'serial_number': row[2].value,
                            'model': row[3].value,
                            'owner_company': row[5].value,
                            'mac_address_ethernet': row[8].value,
                            'status': row[13].value  # Added status field
                        }
                        
                        # Validate required fields
                        missing_fields = [field for field, value in required_fields.items() if not value]
                        if missing_fields:
                            messages.error(request, f'Row {row[0].row}: Missing required fields: {", ".join(missing_fields)}')
                            continue

                        # Get or create asset type
                        asset_type_name = row[1].value
                        try:
                            asset_type = AssetType.objects.get(name__iexact=asset_type_name)
                        except AssetType.DoesNotExist:
                            messages.error(request, f'Row {row[0].row}: Asset Type "{asset_type_name}" not found. Please create it first.')
                            continue

                        # Get the owner company
                        owner_company_name = row[5].value
                        try:
                            owner_company = OwnerCompany.objects.get(name__iexact=owner_company_name)
                        except OwnerCompany.DoesNotExist:
                            messages.error(request, f'Row {row[0].row}: Owner Company "{owner_company_name}" not found. Please create it first.')
                            continue

                        # Validate status
                        status = row[13].value
                        if status not in dict(ITAsset.STATUS_CHOICES):
                            messages.error(request, f'Row {row[0].row}: Invalid status "{status}". Please use one of the available statuses.')
                            continue

                        # Get the assigned employee if provided
                        assigned_employee_id = row[6].value
                        assigned_employee = None
                        if assigned_employee_id:
                            try:
                                assigned_employee = Employee.objects.get(employee_id=assigned_employee_id)
                                # Verify employee belongs to the owner company
                                if assigned_employee.company != owner_company:
                                    messages.error(request, f'Row {row[0].row}: Employee {assigned_employee_id} does not belong to {owner_company_name}. Please assign an employee from the correct company.')
                                    continue
                            except Employee.DoesNotExist:
                                messages.error(request, f'Row {row[0].row}: Employee with ID {assigned_employee_id} not found. Please use a valid employee ID.')
                                continue
                        
                        # Create asset
                        asset = ITAsset(
                            name=row[0].value,
                            asset_type=asset_type,
                            serial_number=row[2].value,
                            model=row[3].value,
                            manufacturer=row[4].value,
                            owner=owner_company,
                            assigned_to=assigned_employee,
                            mac_address_wifi=row[7].value,
                            mac_address_ethernet=row[8].value,
                            delivery_letter_code=row[9].value,
                            status=status,  # Use the provided status
                            purchase_date=row[10].value,
                            receipt_date=row[11].value,
                            warranty_expiry=row[12].value
                        )
                        asset.save()
                        messages.success(request, f'Row {row[0].row}: Asset "{asset.name}" created successfully.')
                    except Exception as e:
                        messages.error(request, f'Row {row[0].row}: Error processing row: {str(e)}')
                        continue
                
                return redirect('inventory:asset_list')
            except Exception as e:
                messages.error(request, f'Error uploading file: {str(e)}')
    
    return render(request, 'inventory/asset_upload.html')

@login_required
def download_asset_data(request):
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        'Asset Name',
        'Asset Type',
        'Serial Number',
        'Model',
        'Manufacturer',
        'Owner',
        'Wi-Fi MAC Address',
        'Ethernet MAC Address',
        'Delivery Letter Code',
        'Status',
        'Purchase Date',
        'Receipt Date',
        'Warranty Expiry'
    ]
    ws.append(headers)
    
    # Get filtered queryset using the same filters as the list view
    assets = ITAsset.objects.all()
    
    # Apply search filter
    search_query = request.GET.get('search', '')
    if search_query:
        assets = assets.filter(
            Q(name__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(model__icontains=search_query) |
            Q(delivery_letter_code__icontains=search_query)
        )
    
    # Apply other filters
    asset_type = request.GET.get('asset_type', '')
    if asset_type:
        assets = assets.filter(asset_type_id=asset_type)
    
    status = request.GET.get('status', '')
    if status:
        assets = assets.filter(status=status)
    
    manufacturer = request.GET.get('manufacturer', '')
    if manufacturer:
        assets = assets.filter(manufacturer=manufacturer)
    
    # Add asset data
    for asset in assets:
        row = [
            asset.name,
            asset.asset_type.name,  # Use the internal name for consistency
            asset.serial_number,
            asset.model,
            asset.manufacturer,
            f"{asset.assigned_to.get_full_name()} ({asset.assigned_to.company})" if asset.assigned_to else '',
            asset.mac_address_wifi,
            asset.mac_address_ethernet,
            asset.delivery_letter_code,
            asset.status,  # Use the internal status value for consistency
            asset.purchase_date,
            asset.receipt_date,
            asset.warranty_expiry
        ]
        ws.append(row)
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=assets.xlsx'
    
    wb.save(response)
    return response

# Asset Type Views
class AssetTypeListView(LoginRequiredMixin, ListView):
    model = AssetType
    template_name = 'inventory/assettype_list.html'
    context_object_name = 'asset_types'

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_assettype'):
            return super().get_queryset()
        return AssetType.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        access_denied = not self.request.user.has_perm('inventory.view_assettype')
        context['access_denied'] = access_denied
        context['can_add_assettype'] = self.request.user.has_perm('inventory.add_assettype') and not access_denied
        context['can_change_assettype'] = self.request.user.has_perm('inventory.change_assettype') and not access_denied
        context['can_delete_assettype'] = self.request.user.has_perm('inventory.delete_assettype') and not access_denied
        return context

class AssetTypeCreateView(SuperuserRequiredMixin, CreateView):
    model = AssetType
    form_class = AssetTypeForm
    template_name = 'inventory/assettype_form.html'
    success_url = reverse_lazy('inventory:assettype_list')

    def form_valid(self, form):
        messages.success(self.request, 'Asset Type created successfully.')
        return super().form_valid(form)

class AssetTypeUpdateView(SuperuserRequiredMixin, UpdateView):
    model = AssetType
    form_class = AssetTypeForm
    template_name = 'inventory/assettype_form.html'
    success_url = reverse_lazy('inventory:assettype_list')

    def form_valid(self, form):
        messages.success(self.request, 'Asset Type updated successfully.')
        return super().form_valid(form)

class AssetTypeDeleteView(SuperuserRequiredMixin, DeleteView):
    model = AssetType
    template_name = 'inventory/assettype_confirm_delete.html'
    success_url = reverse_lazy('inventory:assettype_list')

# Owner Company Views
class OwnerCompanyListView(LoginRequiredMixin, ListView):
    model = OwnerCompany
    template_name = 'inventory/company_list.html'
    context_object_name = 'companies'

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_company'):
            return super().get_queryset()
        return OwnerCompany.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        access_denied = not self.request.user.has_perm('inventory.view_company')
        context['access_denied'] = access_denied
        context['can_add_company'] = self.request.user.has_perm('inventory.add_company') and not access_denied
        context['can_change_company'] = self.request.user.has_perm('inventory.change_company') and not access_denied
        context['can_delete_company'] = self.request.user.has_perm('inventory.delete_company') and not access_denied
        return context

class OwnerCompanyCreateView(SuperuserRequiredMixin, CreateView):
    model = OwnerCompany
    template_name = 'inventory/company_form.html'
    fields = ['name']
    success_url = reverse_lazy('inventory:owner_company_list')

    def form_valid(self, form):
        # Generate code from name: lowercase and replace multiple spaces with single underscore
        name = form.cleaned_data['name']
        code = '_'.join(name.lower().split())
        
        # Limit code to 20 characters
        if len(code) > 20:
            code = code[:20]
        
        # Set the code before saving
        form.instance.code = code
        
        messages.success(self.request, 'Owner Company created successfully.')
        return super().form_valid(form)

class OwnerCompanyUpdateView(SuperuserRequiredMixin, UpdateView):
    model = OwnerCompany
    template_name = 'inventory/company_form.html'
    fields = ['name']
    success_url = reverse_lazy('inventory:owner_company_list')

    def form_valid(self, form):
        # Generate code from name: lowercase and replace multiple spaces with single underscore
        name = form.cleaned_data['name']
        code = '_'.join(name.lower().split())
        
        # Limit code to 20 characters
        if len(code) > 20:
            code = code[:20]
        
        # Set the code before saving
        form.instance.code = code
        
        messages.success(self.request, 'Owner Company updated successfully.')
        return super().form_valid(form)

class OwnerCompanyDeleteView(SuperuserRequiredMixin, DeleteView):
    model = OwnerCompany
    template_name = 'inventory/company_confirm_delete.html'
    success_url = reverse_lazy('inventory:owner_company_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Owner Company deleted successfully.')
        return super().delete(request, *args, **kwargs)

# Department Views
class DepartmentListView(LoginRequiredMixin, ListView):
    model = Department
    template_name = 'inventory/department_list.html'
    context_object_name = 'departments'

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_department'):
            return super().get_queryset()
        return Department.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        access_denied = not self.request.user.has_perm('inventory.view_department')
        context['access_denied'] = access_denied
        context['can_add_department'] = self.request.user.has_perm('inventory.add_department') and not access_denied
        context['can_change_department'] = self.request.user.has_perm('inventory.change_department') and not access_denied
        context['can_delete_department'] = self.request.user.has_perm('inventory.delete_department') and not access_denied
        return context

class DepartmentCreateView(SuperuserRequiredMixin, CreateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'inventory/department_form.html'
    success_url = reverse_lazy('inventory:department_list')

    def form_valid(self, form):
        messages.success(self.request, 'Department created successfully.')
        return super().form_valid(form)

class DepartmentUpdateView(SuperuserRequiredMixin, UpdateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'inventory/department_form.html'
    success_url = reverse_lazy('inventory:department_list')

    def form_valid(self, form):
        messages.success(self.request, 'Department updated successfully.')
        return super().form_valid(form)

class DepartmentDeleteView(SuperuserRequiredMixin, DeleteView):
    model = Department
    template_name = 'inventory/department_confirm_delete.html'
    success_url = reverse_lazy('inventory:department_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Department deleted successfully.')
        return super().delete(request, *args, **kwargs)

class AssetHistoryListView(LoginRequiredMixin, ListView):
    model = AssetHistory
    template_name = 'inventory/asset_history.html'
    context_object_name = 'history_entries'
    paginate_by = 20

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_assethistory'):
            return super().get_queryset().select_related('asset', 'assigned_to')
        return AssetHistory.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = ITAsset.STATUS_CHOICES
        access_denied = not self.request.user.has_perm('inventory.view_assethistory')
        context['access_denied'] = access_denied
        return context

# Team Views
class TeamListView(LoginRequiredMixin, ListView):
    model = Team
    template_name = 'inventory/team_list.html'
    context_object_name = 'teams'

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_team'):
            return super().get_queryset()
        return Team.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        access_denied = not self.request.user.has_perm('inventory.view_team')
        context['access_denied'] = access_denied
        context['can_add_team'] = self.request.user.has_perm('inventory.add_team') and not access_denied
        context['can_change_team'] = self.request.user.has_perm('inventory.change_team') and not access_denied
        context['can_delete_team'] = self.request.user.has_perm('inventory.delete_team') and not access_denied
        return context

class TeamCreateView(SuperuserRequiredMixin, CreateView):
    model = Team
    form_class = TeamForm
    template_name = 'inventory/team_form.html'
    success_url = reverse_lazy('inventory:team_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Only show employed employees in the manager queryset
        form.fields['manager'].queryset = Employee.objects.filter(
            employment_status=Employee.STATUS_EMPLOYED
        )
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get all employed employees except the selected manager
        manager_id = self.request.POST.get('manager')
        available_employees = Employee.objects.filter(
            employment_status=Employee.STATUS_EMPLOYED
        )
        if manager_id:
            available_employees = available_employees.exclude(id=manager_id)
        context['available_employees'] = available_employees
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        team = form.instance
        
        # Get selected team members from form data
        selected_members = self.request.POST.getlist('team_members')
        
        # Add team members
        for member_id in selected_members:
            if member_id != str(team.manager.id):  # Don't add manager as a member
                TeamMember.objects.create(
                    team=team,
                    employee_id=member_id
                )
        
        return response

class TeamUpdateView(SuperuserRequiredMixin, UpdateView):
    model = Team
    form_class = TeamForm
    template_name = 'inventory/team_form.html'
    success_url = reverse_lazy('inventory:team_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Only show employed employees in the manager queryset
        form.fields['manager'].queryset = Employee.objects.filter(
            employment_status=Employee.STATUS_EMPLOYED
        )
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        team = self.get_object()
        
        # Get all employed employees except the current manager
        available_employees = Employee.objects.filter(
            employment_status=Employee.STATUS_EMPLOYED
        ).exclude(id=team.manager.id)
        
        context['available_employees'] = available_employees
        context['current_members'] = team.members.all()
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        team = form.instance
        
        # Get selected team members from form data
        selected_members = self.request.POST.getlist('team_members')
        
        # Remove all existing team members
        team.members.all().delete()
        
        # Add new team members
        for member_id in selected_members:
            if member_id != str(team.manager.id):  # Don't add manager as a member
                TeamMember.objects.create(
                    team=team,
                    employee_id=member_id
                )
        
        return response

class TeamDeleteView(SuperuserRequiredMixin, DeleteView):
    model = Team
    template_name = 'inventory/team_confirm_delete.html'
    success_url = reverse_lazy('inventory:team_list')

    def delete(self, request, *args, **kwargs):
        team = self.get_object()
        # Update all team members to remove team association
        Employee.objects.filter(team=team).update(team=None)
        return super().delete(request, *args, **kwargs)

class TeamDetailView(LoginRequiredMixin, DetailView):
    model = Team
    template_name = 'inventory/team_detail.html'
    context_object_name = 'team'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['members'] = self.object.members.all().select_related('department', 'company')
        context['available_employees'] = Employee.objects.filter(
            employment_status=Employee.STATUS_EMPLOYED,
            department=self.object.department,
            team__isnull=True
        ).select_related('department', 'company')
        
        # Check if user has an employee profile before accessing it
        try:
            context['is_manager'] = self.object.manager == self.request.user.employee_profile
        except:
            context['is_manager'] = False
            
        return context

@login_required
def add_team_member(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    
    # Check if user is the team manager
    if request.user.employee_profile != team.manager:
        messages.error(request, 'Only the team manager can add members.')
        return redirect('inventory:team_detail', pk=team_id)
    
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        if employee_id:
            employee = get_object_or_404(Employee, id=employee_id)
            if employee.department == team.department:
                employee.team = team
                employee.save()
                messages.success(request, f'{employee.get_full_name()} has been added to the team.')
            else:
                messages.error(request, 'Employee must be from the same department as the team.')
    return redirect('inventory:team_detail', pk=team_id)

@login_required
def remove_team_member(request, team_id, employee_id):
    team = get_object_or_404(Team, id=team_id)
    
    # Check if user is the team manager
    if request.user.employee_profile != team.manager:
        messages.error(request, 'Only the team manager can remove members.')
        return redirect('inventory:team_detail', pk=team_id)
    
    employee = get_object_or_404(Employee, id=employee_id)
    if employee.team == team:
        employee.team = None
        employee.save()
        messages.success(request, f'{employee.get_full_name()} has been removed from the team.')
    return redirect('inventory:team_detail', pk=team_id)

@login_required
def my_teams(request):
    """View for team managers to see their teams."""
    managed_teams = Team.objects.filter(manager=request.user.employee_profile)
    return render(request, 'inventory/my_teams.html', {
        'teams': managed_teams
    })

# User Management Views
class UserListView(SuperuserRequiredMixin, ListView):
    model = User
    template_name = 'inventory/user_list.html'
    context_object_name = 'users'

class UserCreateView(SuperuserRequiredMixin, CreateView):
    model = User
    form_class = UserForm
    template_name = 'inventory/user_form.html'
    success_url = reverse_lazy('inventory:user_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['profile_form'] = UserProfileForm(self.request.POST)
        else:
            context['profile_form'] = UserProfileForm()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        profile_form = context['profile_form']
        if profile_form.is_valid():
            user = form.save(commit=False)
            user.is_active = True
            user.save()
            
            # Get or create profile
            profile, created = UserProfile.objects.get_or_create(
                user=user,
                defaults={'role': None}  # Don't set default role, we'll set it from form
            )
            
            # Set role from form
            role = profile_form.cleaned_data.get('role')
            if role:
                profile.role = role
                # Clear any existing permissions and set only the role's permissions
                user.user_permissions.clear()
                for permission in role.permissions.all():
                    user.user_permissions.add(permission)
            else:
                # If no role selected, set default role but don't grant any permissions
                default_role = Role.objects.first()
                if default_role:
                    profile.role = default_role
                    # Don't grant any permissions for default role
                    user.user_permissions.clear()
                else:
                    messages.error(self.request, 'No roles available in the system. Please create a role first.')
                    return self.form_invalid(form)

            # Update other profile fields
            profile.department = profile_form.cleaned_data.get('department')
            profile.team = profile_form.cleaned_data.get('team')
            profile.save()

            # Assign user to employee if selected
            employee = profile_form.cleaned_data.get('employee')
            if employee:
                employee.user = user
                employee.save()

            messages.success(self.request, 'User created successfully.')
            return redirect(self.success_url)
        return self.render_to_response(self.get_context_data(form=form))

class UserUpdateView(SuperuserRequiredMixin, UpdateView):
    model = User
    form_class = UserForm
    template_name = 'inventory/user_form.html'
    success_url = reverse_lazy('inventory:user_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['profile_form'] = UserProfileForm(self.request.POST, instance=self.object.profile)
        else:
            context['profile_form'] = UserProfileForm(instance=self.object.profile)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        profile_form = context['profile_form']
        if profile_form.is_valid():
            user = form.save()
            profile_form.save()

            newly_assigned_employee = profile_form.cleaned_data.get('employee')
            currently_assigned_employee = user.employee_profile if hasattr(user, 'employee_profile') else None

            if newly_assigned_employee != currently_assigned_employee:
                # Un-assign the old employee if there was one
                if currently_assigned_employee:
                    currently_assigned_employee.user = None
                    currently_assigned_employee.save()
                
                # Assign the new employee if one was selected
                if newly_assigned_employee:
                    newly_assigned_employee.user = user
                    newly_assigned_employee.save()

            messages.success(self.request, 'User updated successfully.')
            return redirect(self.success_url)
        return self.render_to_response(self.get_context_data(form=form))

class UserDeleteView(SuperuserRequiredMixin, DeleteView):
    model = User
    template_name = 'inventory/user_confirm_delete.html'
    success_url = reverse_lazy('inventory:user_list')

    def get(self, request, *args, **kwargs):
        # Check if this is the last user
        if User.objects.count() <= 1:
            messages.error(request, 'Cannot delete the last user in the system.')
            return redirect('inventory:user_list')

        # Get the user to be deleted
        user_to_delete = self.get_object()
        
        # Check if trying to delete an admin user
        try:
            if user_to_delete.profile and user_to_delete.profile.role and user_to_delete.profile.role.name == 'Administrator':
                # Only allow deletion if the current user is also an admin
                if not request.user.profile or not request.user.profile.role or request.user.profile.role.name != 'Administrator':
                    messages.error(request, 'Only administrators can delete administrator users.')
                    return redirect('inventory:user_list')
        except:
            # If user has no profile, allow deletion
            pass

        return super().get(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        # Check if this is the last user
        if User.objects.count() <= 1:
            messages.error(request, 'Cannot delete the last user in the system.')
            return redirect('inventory:user_list')

        # Get the user to be deleted
        user_to_delete = self.get_object()
        
        # Check if trying to delete an admin user
        try:
            if user_to_delete.profile and user_to_delete.profile.role and user_to_delete.profile.role.name == 'Administrator':
                # Only allow deletion if the current user is also an admin
                if not request.user.profile or not request.user.profile.role or request.user.profile.role.name != 'Administrator':
                    messages.error(request, 'Only administrators can delete administrator users.')
                    return redirect('inventory:user_list')
        except:
            # If user has no profile, allow deletion
            pass
            
        messages.success(request, 'User deleted successfully.')
        return super().delete(request, *args, **kwargs)

# Role Management Views
class RoleListView(LoginRequiredMixin, ListView):
    model = Role
    template_name = 'inventory/role_list.html'
    context_object_name = 'roles'

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_role'):
            return super().get_queryset()
        return Role.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['access_denied'] = not self.request.user.has_perm('inventory.view_role')
        return context

class RoleCreateView(SuperuserRequiredMixin, CreateView):
    model = Role
    form_class = RoleForm
    template_name = 'inventory/role_form.html'
    success_url = reverse_lazy('inventory:role_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass the full permission set to the template for rendering checkboxes
        all_permissions = Role.get_all_defined_permissions()
        context['permissions'] = {
            category: list(perms.items())
            for category, perms in all_permissions.items()
        }
        return context

    def form_valid(self, form):
        # Manually process the form and permissions to prevent errors.
        # This gives us full control over the instance creation and redirection.
        self.object = form.save(commit=False)
        
        permissions = {}
        permission_values = self.request.POST.getlist('permissions')
        
        for perm in permission_values:
            permissions[perm] = True
        
        self.object.permissions = permissions
        self.object.save()
        
        messages.success(self.request, 'Role created successfully.')
        return redirect(self.get_success_url())

class RoleUpdateView(SuperuserRequiredMixin, UpdateView):
    model = Role
    form_class = RoleForm
    template_name = 'inventory/role_form.html' # Use the unified template
    success_url = reverse_lazy('inventory:role_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get permissions from the single source of truth in the Role model
        all_permissions = Role.get_all_defined_permissions()
        
        # The template expects a list of tuples for iteration, so we convert the inner dicts
        context['permissions'] = {
            category: list(perms.items())
            for category, perms in all_permissions.items()
        }
        
        return context

    def form_valid(self, form):
        # Manually process the form and permissions to ensure they are saved correctly.
        self.object = form.save(commit=False)
        
        permissions = {}
        permission_values = self.request.POST.getlist('permissions')
        
        for perm in permission_values:
            permissions[perm] = True
        
        self.object.permissions = permissions
        self.object.save()
        
        messages.success(self.request, 'Role updated successfully.')
        return redirect(self.get_success_url())

class RoleDeleteView(SuperuserRequiredMixin, DeleteView):
    model = Role
    template_name = 'inventory/role_confirm_delete.html'
    success_url = reverse_lazy('inventory:role_list')

    def get(self, request, *args, **kwargs):
        role = self.get_object()
        
        # Prevent deletion of Administrator role
        if role.name == 'Administrator':
            messages.error(request, 'Cannot delete the Administrator role.')
            return redirect('inventory:role_list')
            
        return super().get(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        role = self.get_object()
        
        # Prevent deletion of Administrator role
        if role.name == 'Administrator':
            messages.error(request, 'Cannot delete the Administrator role.')
            return redirect('inventory:role_list')
            
        messages.success(request, 'Role deleted successfully.')
        return super().delete(request, *args, **kwargs)

# Authentication Views
@login_required
def logout_view(request):
    logout(request)
    return redirect('/accounts/login/')

# Ticket Views
class TicketListView(LoginRequiredMixin, ListView):
    model = Ticket
    template_name = 'inventory/ticket_list.html'
    context_object_name = 'tickets'

    def get_queryset(self):
        if self.request.user.has_perm('inventory.view_ticket'):
            return Ticket.objects.filter(
                Q(sender=self.request.user) | Q(recipient=self.request.user),
                parent_ticket__isnull=True
            ).select_related('sender', 'recipient', 'asset', 'employee').order_by('-created_at')
        return Ticket.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        access_denied = not self.request.user.has_perm('inventory.view_ticket')
        context['access_denied'] = access_denied
        context['can_add_ticket'] = self.request.user.has_perm('inventory.add_ticket') and not access_denied
        context['can_change_ticket'] = self.request.user.has_perm('inventory.change_ticket') and not access_denied
        context['can_delete_ticket'] = self.request.user.has_perm('inventory.delete_ticket') and not access_denied
        context['unread_tickets_count'] = Ticket.objects.filter(
            recipient=self.request.user,
            is_read=False
        ).count() if not access_denied else 0
        context['has_new_replies'] = Ticket.objects.filter(
            recipient=self.request.user,
            replies__has_new_reply=True
        ).exists() if not access_denied else False
        return context

class TicketCreateView(SuperuserRequiredMixin, CreateView):
    model = Ticket
    form_class = TicketForm
    template_name = 'inventory/ticket_form.html'
    success_url = reverse_lazy('inventory:ticket_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['users'] = User.objects.filter(is_active=True).exclude(id=self.request.user.id).order_by('username')
        context['unread_tickets_count'] = Ticket.objects.filter(
            recipient=self.request.user,
            is_read=False
        ).count()
        return context

    def form_valid(self, form):
        try:
            recipient_id = self.request.POST.get('recipient')
            if not recipient_id:
                messages.error(self.request, 'Please select a recipient.')
                return self.form_invalid(form)
            
            recipient = get_object_or_404(User, id=recipient_id)
            ticket = form.save(commit=False)
            ticket.sender = self.request.user
            ticket.recipient = recipient
            ticket.save()
            messages.success(self.request, 'Ticket created successfully.')
            return redirect(self.success_url)
        except Exception as e:
            messages.error(self.request, f'Error creating ticket: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'There was an error creating your ticket. Please check the form and try again.')
        return self.render_to_response(self.get_context_data(form=form))

class TicketDetailView(SuperuserRequiredMixin, DetailView):
    model = Ticket
    template_name = 'inventory/ticket_detail.html'
    context_object_name = 'ticket'

    def get_queryset(self):
        return Ticket.objects.filter(
            Q(sender=self.request.user) | Q(recipient=self.request.user)
        ).select_related('sender', 'recipient', 'asset', 'employee')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ticket = self.get_object()
        
        # Mark ticket as read if the current user is the recipient
        if self.request.user == ticket.recipient and not ticket.is_read:
            ticket.mark_as_read()
        
        # Get all replies to this ticket
        replies = ticket.get_replies()
        
        # Mark replies as read and clear has_new_reply flag when viewing the ticket
        if self.request.user == ticket.recipient:
            Ticket.objects.filter(
                parent_ticket=ticket,
                recipient=self.request.user,
                is_read=False
            ).update(
                is_read=True,
                has_new_reply=False
            )
            
            ticket.has_new_reply = False
            ticket.save(update_fields=['has_new_reply'])
        
        context['replies'] = replies
        context['unread_tickets_count'] = Ticket.objects.filter(
            recipient=self.request.user,
            is_read=False
        ).count()
        return context

    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        if self.object.recipient == request.user and not self.object.is_read:
            Ticket.objects.filter(pk=self.object.pk).update(is_read=True)
            self.object.refresh_from_db()
        return response

class TicketDeleteView(SuperuserRequiredMixin, DeleteView):
    model = Ticket
    template_name = 'inventory/ticket_confirm_delete.html'
    success_url = reverse_lazy('inventory:ticket_list')

    def get_queryset(self):
        return Ticket.objects.filter(
            Q(sender=self.request.user) | Q(recipient=self.request.user)
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_tickets_count'] = Ticket.objects.filter(
            recipient=self.request.user,
            is_read=False
        ).count()
        context['source_url'] = self.request.GET.get('source', 'inbox')
        return context

    def get_success_url(self):
        source = self.request.GET.get('source', 'inbox')
        
        if source == 'sent':
            return reverse_lazy('inventory:ticket_list')
        elif source == 'detail':
            ticket = self.get_object()
            if ticket.parent_ticket:
                return reverse_lazy('inventory:ticket_detail', kwargs={'pk': ticket.parent_ticket.id})
            return reverse_lazy('inventory:ticket_list')
        return reverse_lazy('inventory:ticket_list')

    def delete(self, request, *args, **kwargs):
        ticket = self.get_object()
        
        if ticket.sender != request.user and ticket.recipient != request.user:
            messages.error(request, 'You can only delete your own tickets.')
            return redirect('inventory:ticket_list')
        
        ticket.delete()
        messages.success(request, 'Ticket deleted successfully.')
        return redirect(self.get_success_url())

class TicketReplyView(SuperuserRequiredMixin, CreateView):
    model = Ticket
    form_class = TicketForm
    template_name = 'inventory/ticket_reply.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        original_ticket = get_object_or_404(Ticket, pk=self.kwargs['pk'])
        context['original_ticket'] = original_ticket
        context['unread_tickets_count'] = Ticket.objects.filter(
            recipient=self.request.user,
            is_read=False
        ).count()
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        original_ticket = get_object_or_404(Ticket, pk=self.kwargs['pk'])
        kwargs['original_message'] = original_ticket
        return kwargs

    def form_valid(self, form):
        try:
            original_ticket = get_object_or_404(Ticket, pk=self.kwargs['pk'])
            ticket = form.save(commit=False)
            ticket.sender = self.request.user
            ticket.recipient = original_ticket.sender
            ticket.parent_ticket = original_ticket
            ticket.has_new_reply = True
            ticket.save()

            original_ticket.has_new_reply = True
            original_ticket.save(update_fields=['has_new_reply'])

            return redirect('inventory:ticket_detail', pk=original_ticket.id)
        except Exception as e:
            messages.error(self.request, f'Error sending reply: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'There was an error sending your reply. Please check the form and try again.')
        return self.render_to_response(self.get_context_data(form=form))

def ticket_notifications(request):
    """Stream ticket notifications using Server-Sent Events."""
    def event_stream():
        last_check = time.time()
        while True:
            new_tickets = Ticket.objects.filter(
                recipient=request.user,
                created_at__gt=time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(last_check))
            ).exists()
            
            if new_tickets:
                data = json.dumps({'has_new_tickets': True})
                yield f"data: {data}\n\n"
            
            last_check = time.time()
            time.sleep(5)
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response

@login_required
def check_replies(request, pk):
    """Check if there are new replies to a ticket."""
    ticket = get_object_or_404(Ticket, pk=pk)
    
    last_check = request.GET.get('last_check')
    if last_check:
        try:
            last_check = datetime.fromisoformat(last_check.replace('Z', '+00:00'))
            has_new_replies = ticket.replies.filter(created_at__gt=last_check).exists()
        except (ValueError, TypeError):
            has_new_replies = False
    else:
        has_new_replies = False
    
    return JsonResponse({'has_new_replies': has_new_replies})

def ticket_context_processor(request):
    """Add ticket-related context to all templates."""
    if request.user.is_authenticated:
        return {
            'unread_tickets_count': Ticket.objects.filter(
                recipient=request.user,
                is_read=False
            ).count(),
            'has_new_replies': Ticket.objects.filter(
                recipient=request.user,
                replies__has_new_reply=True
            ).exists()
        }
    return {
        'unread_tickets_count': 0,
        'has_new_replies': False
    }

def get_department_positions(request, department_id):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    positions = Position.objects.filter(department_id=department_id).order_by('name')
    positions_data = [{'id': pos.id, 'name': pos.name} for pos in positions]
    return JsonResponse(positions_data, safe=False)

def custom_403(request, exception=None):
    return render(request, '403.html', status=403)

def custom_404(request, exception=None):
    return render(request, '404.html', status=404)  